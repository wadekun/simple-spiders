# _*_ encoding:utf-8 _*_

import urllib
import urllib2
import cookielib
import os
import sys
import re
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf-8')


class Kstore_Order_Spider(object):
    def __init__(self):
        self.__login_url = 'http://kstore.qianmi.com/checklogin.htm'
        self.__order_url = 'http://kstore.qianmi.com/myorder.htm?pageNo='
        self.__cookie = cookielib.CookieJar()
        self.__opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(self.__cookie))
        self.__order_code_pattern = re.compile(r'\d{16}')
        self.__order_amount_pattern = re.compile(r'\d+\.\d{2}')
        self.__login_data = urllib.urlencode({'username': '', 'password': '', 'url': '', 'type': '0'})

    def run(self):
        self.login()
        self.write_excel(self.crawl_order())

    def login(self):
        request = urllib2.Request(self.__login_url, self.__login_data)
        response = self.__opener.open(request)
        print response.read()

    def crawl_order(self):
        page_no = 0
        order_list = []
        while True:
            page_no += 1
            request = urllib2.Request(self.__order_url+str(page_no))
            response = self.__opener.open(request)
            html = response.read()
            soup = BeautifulSoup(html)
            order_list_div = soup.find('div', {'class': 'new_order_list'})
            order_body_list = order_list_div.find_all('tbody')
            print '第 %d 页共有 %d 个订单' % (page_no, len(order_body_list))
            for order in order_body_list:
                order_code = self.__get_order_code(order)
                print '订单 %s ====================' % order_code
                order_goods = self.__get_order_goods(order)
                customer_name = self.__get_customer(order)
                order_amount = self.__get_order_amount(order)
                print '=========================\n'
                order_list.append([order_code, order_goods, customer_name, order_amount])
            if not self.__has_next_page(order_list_div):
                break
        return order_list

    def __has_next_page(self, order_content):
        non_next_btn = order_content.find('div', {'class': 'paging_area'}).find('a', {'class': 'next_null'})
        # print non_next_btn
        if non_next_btn:
            return False
        else:
            return True

    def write_excel(self, order_list):
        wb = Workbook()
        ws = wb.active
        ws = wb.create_sheet()
        ws.title = 'customer orders'
        ws.append(['订单号', '订单商品', '会员名', '订单金额'])
        for order in order_list:
            ws.append(order)
        save_path = 'order_list.' + str(time.time()) + '.xlsx'
        print save_path
        wb.save(save_path)

    def __get_order_amount(self, order):
        order_content = order.find('tr', {'class': 'order-bd'})
        amount_text = order_content.find_all('td')[3].get_text()
        amount_match = self.__order_amount_pattern.search(amount_text)
        if amount_match:
            print amount_match.group()
            return amount_match.group()
        else:
            return 00.00


    def __get_customer(self, order):
        order_content = order.find('tr', {'class': 'order-bd'})
        customer = order_content.find_all('td')[1].get_text()
        name = ' '.join(customer.split())
        print name
        return name


    def __get_order_goods(self, order):
        order_content = order.find('tr', {'class': 'order-bd'})
        order_goods_descs = order_content.find('td', {'class': 'baobei'})\
        .find_all('div', {'class': 'desc'})
        order_goods = []
        for goods_desc in order_goods_descs:
            goods_title = goods_desc.find('a', {'class': 'name'}).get_text()
            print goods_title
            order_goods.append(goods_title)
        return '\n'.join(order_goods)


    def __get_order_code(self, order):
        order_header = order.find('tr', {'class': 'order-hd'})
        order_code_text = order_header.find('td', {'class': 'first'}).find('span').get_text()
        code_match = self.__order_code_pattern.search(order_code_text)
        if code_match:
            return code_match.group()
        else:
            return ''


def test_re():
    text = '胜多负少12345678'
    pattern = re.compile(r'\d+')
    match = pattern.search(text)
    if match:
        print match.group()
    else:
        print '不匹配'


def test_write_excel():
    wb = Workbook()
    ws = wb.active             #默认创建第一个表，默认名字为sheet
    ws1 = wb.create_sheet()    #创建第二个表
    ws1.title = "New Title"    #为第二个表设置名字
    ws2 = wb.get_sheet_by_name("New Title")                #通过名字获取表，和第二个表示一个表
    wb.save('your_name.xlsx') #保存


if __name__ == '__main__':
    # test_re()
    # test_write_excel()
    spider = Kstore_Order_Spider()
    spider.run()
    # spider.login()
